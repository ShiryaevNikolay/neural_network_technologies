from openpyxl import *


def read_weights(length, file_name='../weights.xlsx', sheet_name='weights'):
    """
    Считывает веса из файла таблицы
    :param length: Размер данных в одной строке/столбце
    :param file_name: Имя файла
    :param sheet_name: Название страницы с весами
    :return: Сдвиг/нулевой вес и матрицу весов размером length
    """
    weights_file = load_workbook(file_name)
    weights = []
    shift = 0.5
    try:
        sheet = weights_file.get_sheet_by_name(sheet_name)
        shift = sheet.cell(column=1, row=1).value
        for i in range(length):
            weights.append([])
            for j in range(length):
                weight_value = sheet.cell(column=i + 2, row=j + 1).value
                weights[i].append(weight_value)
        weights_file.save(file_name)
        weights_file.close()
        return shift, weights
    except:
        weights_file.save(file_name)
        weights_file.close()
        return shift, weights
