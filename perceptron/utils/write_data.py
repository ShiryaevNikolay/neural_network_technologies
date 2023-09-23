from openpyxl import *


def save_weights(shift, weights, sheet_name='weights', file_name='weights.xlsl'):
    """
    Сохраняет веса в файл
    :param shift: Сдвиг/смещение/нулевой вес
    :param weights: Матрица с весами
    :param sheet_name: Название страницы в таблице, где будут сохранены веса
    :param file_name: Название файла
    """
    weights_file = load_workbook(file_name)
    try:
        sheet = weights_file.get_sheet_by_name(sheet_name)
    except:
        weights_file.create_sheet(sheet_name)
        sheet = weights_file.get_sheet_by_name(sheet_name)
    sheet.cell(column=1, row=1).value = shift
    column_length = len(weights)
    for i in range(column_length):
        row_length = len(weights[i])
        for j in range(row_length):
            sheet.cell(column=i + 2, row=j + 1).value = weights[i][j]
    weights_file.save(file_name)
    weights_file.close()
